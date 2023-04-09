import json
from django.shortcuts import render
import os
from config import settings
from django.http import JsonResponse
from  common.models import Task
from openpyxl import load_workbook

# Create your views here.

def submit(request):
    # print(request)
    name = request.POST.get("name")
    id = request.POST.get("id")
    img = request.FILES.get("img")

    if name and id and img:
        
        # 从数据库读出要存入的路径
        qs = Task.objects.values()
        qs = qs.filter(id=id)
        filepath = list(qs)[0]['filePath']
        # print(filepath)

        # 读取excel文件，并将提交结果写入
        wb = load_workbook(filepath + "\\提交名单.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        minrow=sheet.min_row #最小行 从1开始
        maxrow=sheet.max_row #最大行
        # i 行 j 列
        for row in range(minrow, maxrow + 1):
            # 第一列的名字用于判断
            cell=sheet.cell(row, 1).value
            if row != maxrow: 
                if cell == name:
                    sheet['C' + str(row)] = 'yes'
                    wb.save(filepath + "\\提交名单.xlsx")
                    break;
            else:
                data = {"ret":"1","msg":"你不在本次任务名单中"}
                return JsonResponse(data)
        
        # 图片重命名
        last_name = img.name.split(".")[-1]  # 后缀名
        img.name = name + "." + last_name
        # print(img.name) 

        # 将图片存放到路径下
        destination = open(os.path.join(filepath, img.name),'wb+')
        for chunk in img.chunks():
            destination.write(chunk)
        destination.close()

        data = {"ret":0,"msg":"上传成功"}
        return JsonResponse(data)
    else:
        data = {"ret":"2","msg":"文件上传失败"}
        return JsonResponse(data)