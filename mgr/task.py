
from django.http import HttpResponse
# 导入 Task 对象定义
from  common.models import Task
# 到入创建文件夹的功能
from .function import mkdir
from django.http import JsonResponse
import json
import os
from config import settings
import shutil
from openpyxl import load_workbook
# Create your views here.

def helloworld(request):
    return HttpResponse("Helloworld")


def list_task(request):
    # 返回一个 QuerySet 对象 ，包含所有的表记录
    # 每条表记录都是是一个dict对象，
    # key 是字段名，value 是 字段值
    qs = Task.objects.values()
    # 转换为字符串
    retlist = list(qs)
    data = {'ret': '0', 'retlist': retlist }
    return JsonResponse(data)


def add_task(request):
    info = request.params['data']

    # 从请求消息中 获取要添加任务的信息
    # 并且插入到数据库中
    # 返回值 就是对应插入记录的对象

    # 创建一个目录用于存放收集的文件
    # 存放文件的目录
    dir = os.path.join(settings.BASE_DIR, 'upload')
    # 人员的目录
    list_name =  dir + "\\excel\\提交名单.xlsx"
    # 新建存放文件夹的目录
    file_name = dir + "\\" + info['title']
    # print()
    # print(list_name)
    # print(file_path)
    # print()
    isSuccess = mkdir(file_name)
    if isSuccess:
        shutil.copyfile(list_name, file_name + "\\提交名单.xlsx")
        record = Task.objects.create(title=info['title'],
                                    desc=info['desc'],
                                    filePath = file_name,
                                    fileContent = info['fileContent'],
        )

        # print(info)
        return JsonResponse({'ret': 0, 'id':record.id})
    else:
        return JsonResponse({'ret': 1, 'msg':"同名任务已存在"})


def delete_task(request):
    taskid = request.params['id']

    try:
        # 根据 id 从数据库中找到相应的任务记录
        task = Task.objects.get(id=taskid)
    except Task.DoesNotExist:
        return  {
                'ret': "1",
                'msg': f'id 为`{taskid}`的任务不存在'
        }

    # delete 方法就将该记录从数据库中删除了
    task.delete()
    return JsonResponse({'ret': 0})


def task_details(request):
    taskid = request.params['id']
    # 从数据库读出文件存放的路径，以及任务标题
    qs = Task.objects.values()
    qs = qs.filter(id=taskid)
    title = list(qs)[0]['title']
    filepath = list(qs)[0]['filePath']

    # 查询提交人员
    submitted = list()
    unsubmitted = list()
    # 读取excel文件，并将提交结果写入
    wb = load_workbook(filepath + "\\提交名单.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    minrow=sheet.min_row #最小行 从1开始
    maxrow=sheet.max_row #最大行
    # i 行 j 列
    for row in range(minrow + 1, maxrow + 1):
        # 第三列的值用于判断是否提交
        cell=sheet.cell(row, 3).value
        if cell == 'yes':
            submitted.append(sheet.cell(row, 1).value)
        else:
            unsubmitted.append(sheet.cell(row, 1).value)
        
    return JsonResponse({"ret": 0, "id":taskid, "title":title, "submitted":submitted, "unsubmitted":unsubmitted})


def dispatcher(request):
    # 根据session判断用户是否是登录的管理员用户
    if 'usertype' not in request.session:
        return JsonResponse({
            'ret': 302,
            'msg': '未登录',
            'redirect': '/login.html'}, 
            status=302)

    if request.session['usertype'] != 'mgr' :
        return JsonResponse({
            'ret': 302,
            'msg': '用户非mgr类型',
            'redirect': '/login.html'},
            status=302)

    # 将请求参数统一放入request 的 params 属性中，方便后续处理

    # GET请求 参数在url中，同过request 对象的 GET属性获取
    if request.method == 'GET':
        request.params = request.GET

    # POST/PUT/DELETE 请求 参数 从 request 对象的 body 属性中获取
    elif request.method in ['POST','PUT','DELETE']:
        
        # 根据接口，POST/PUT/DELETE 请求的消息体都是 json格式
        print(request.body)
        request.params = json.loads(request.body)

    # 根据不同的action分派给不同的函数进行处理
    action = request.params['action']
    if action == 'list_task':
        return list_task(request)
    elif action == 'add_task':
        return add_task(request)
    elif action == 'del_task':
        return delete_task(request)
    elif action == 'task_details':
        return task_details(request)
    else:
        return JsonResponse({'ret': 1, 'msg': '不支持该类型http请求'})
